import json
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.sessions import async_session_maker
from app.models.db_models import Transaction, User
from app.schemas.enums import TransactionStatusEnum, TransactionTypeEnum


async def get_new_users_count(session: AsyncSession, start_date: date, end_date: date) -> int:
    """
    Returns the count of new users created within the date range (inclusive).
    """
    query = (
        select(func.count())
        .select_from(User)
        .where(
                func.date(User.created) >= start_date,
                func.date(User.created) <= end_date,
        )
    )
    return (await session.scalar(query)) or 0


async def get_distinct_senders_count(
    session: AsyncSession,
    start_date: date,
    end_date: date,
    txn_type: Optional[str] = None,
    txn_status: Optional[str] = None,
) -> int:
    """
    Returns the number of unique users (sender_id) in the specified date range.
    You can filter by transaction type and status if provided.
    """
    conditions = [func.date(Transaction.created) >= start_date, func.date(Transaction.created) <= end_date]
    if txn_type:
        conditions.append(Transaction.type == txn_type)
    if txn_status:
        conditions.append(Transaction.status == txn_status)

    query = select(func.count(func.distinct(Transaction.sender_id))).where(*conditions)
    return (await session.scalar(query)) or 0


async def get_transaction_sum(
    session: AsyncSession,
    start_date: date,
    end_date: date,
    txn_type: Optional[str] = None,
    txn_status: Optional[str] = None,
) -> float:
    """
    Returns the total transaction amount within the specified date range.
    Transaction type and status can be specified if needed.
    """
    conditions = [func.date(Transaction.created) >= start_date, func.date(Transaction.created) <= end_date]
    if txn_type:
        conditions.append(Transaction.type == txn_type)
    if txn_status:
        conditions.append(Transaction.status == txn_status)

    query = select(func.coalesce(func.sum(Transaction.amount), 0)).where(*conditions)
    result = await session.scalar(query)
    return float(result or 0)


async def get_transaction_count(
    session: AsyncSession,
    start_date: date,
    end_date: date,
    txn_status: Optional[str] = None,
    txn_type: Optional[str] = None,
) -> int:
    """
    Returns the number of transactions within the specified date range.
    You can filter by status and type if needed.
    """
    conditions = [func.date(Transaction.created) >= start_date, func.date(Transaction.created) <= end_date]
    if txn_status:
        conditions.append(Transaction.status == txn_status)
    if txn_type:
        conditions.append(Transaction.type == txn_type)

    query = select(func.count()).where(*conditions)
    return (await session.scalar(query)) or 0


async def get_average_transaction_amount(
    session: AsyncSession,
    start_date: date,
    end_date: date,
    txn_type: Optional[str] = None,
    txn_status: Optional[str] = None,
) -> float:
    """
    Returns the average transaction amount within the specified date range.
    You can filter by status and type if needed.
    """
    conditions = [func.date(Transaction.created) >= start_date, func.date(Transaction.created) <= end_date]
    if txn_type:
        conditions.append(Transaction.type == txn_type)
    if txn_status:
        conditions.append(Transaction.status == txn_status)

    query = select(func.coalesce(func.avg(Transaction.amount), 0)).where(*conditions)
    result = await session.scalar(query)
    return float(result or 0)


async def get_conversions(session: AsyncSession, start_date: date, end_date: date) -> Dict[str, Any]:
    """
    Returns a dictionary in the following format:
    {
       "USD_to_EUR": {"count": 10, "sum_amount": 1000.0},
       ...
    }
    for transactions of type EXCHANGE with status PROCESSED during the period.
    """
    conditions = [
        func.date(Transaction.created) >= start_date,
        func.date(Transaction.created) <= end_date,
        Transaction.type == TransactionTypeEnum.EXCHANGE.value,
        Transaction.status == TransactionStatusEnum.PROCESSED.value,
    ]

    query = (
        select(
            Transaction.from_currency,
            Transaction.to_currency,
            func.count().label("count"),
            func.coalesce(func.sum(Transaction.amount), 0).label("sum_amount"),
        )
        .where(*conditions)
        .group_by(Transaction.from_currency, Transaction.to_currency)
    )
    results = await session.execute(query)
    conversions = {}
    for row in results:
        key = f"{row.from_currency}_to_{row.to_currency}"
        conversions[key] = {
            "count": row.count,
            "sum_amount": float(row.sum_amount),
        }
    return conversions


async def collect_week_metrics(
    session: AsyncSession,
    week_start: datetime,
    week_end: datetime,
    previous: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """
    Collects weekly metrics from week_start to week_end (inclusive).
    The 'previous' parameter contains data from the previous week (if available)
    and is used to calculate changes.
    Returns a dictionary with the metrics.
    """
    week_start_date = week_start.date()
    week_end_date = week_end.date()

    # Get values using helper functions
    new_users = await get_new_users_count(session, week_start_date, week_end_date)
    deposit_users = await get_distinct_senders_count(
        session, week_start_date, week_end_date, txn_type=TransactionTypeEnum.DEPOSIT
    )
    transaction_users = await get_distinct_senders_count(session, week_start_date, week_end_date)

    sum_deposits = await get_transaction_sum(
        session,
        week_start_date,
        week_end_date,
        txn_type=TransactionTypeEnum.DEPOSIT,
        txn_status=TransactionStatusEnum.PROCESSED,
    )
    sum_withdrawals = await get_transaction_sum(
        session,
        week_start_date,
        week_end_date,
        txn_type=TransactionTypeEnum.WITHDRAWAL,
        txn_status=TransactionStatusEnum.PROCESSED,
    )
    sum_transfers = await get_transaction_sum(
        session,
        week_start_date,
        week_end_date,
        txn_type=TransactionTypeEnum.TRANSFER,
        txn_status=TransactionStatusEnum.PROCESSED,
    )

    total_transactions = await get_transaction_count(session, week_start_date, week_end_date)
    completed_transactions = await get_transaction_count(
        session, week_start_date, week_end_date, txn_status=TransactionStatusEnum.PROCESSED
    )

    conversions = await get_conversions(session, week_start_date, week_end_date)

    avg_deposit = await get_average_transaction_amount(
        session,
        week_start_date,
        week_end_date,
        txn_type=TransactionTypeEnum.DEPOSIT,
        txn_status=TransactionStatusEnum.PROCESSED,
    )
    avg_withdrawal = await get_average_transaction_amount(
        session,
        week_start_date,
        week_end_date,
        txn_type=TransactionTypeEnum.WITHDRAWAL,
        txn_status=TransactionStatusEnum.PROCESSED,
    )

    active_users = await get_distinct_senders_count(session, week_start_date, week_end_date)

    # Function to calculate the difference and percentage change
    def calc_delta(current: float, prev: Optional[float]) -> Dict[str, Optional[float]]:
        if prev is None:
            return {"delta": None, "pct_change": None}
        delta = current - prev
        pct = (delta / prev * 100.0) if prev != 0 else None
        return {"delta": delta, "pct_change": pct}

    # If previous week's data is available, calculate the changes
    dynamics = {}
    if previous:
        dynamics["new_users"] = calc_delta(new_users, previous.get("new_users"))
        dynamics["sum_deposits"] = calc_delta(sum_deposits, previous.get("sum_deposits"))
        dynamics["sum_withdrawals"] = calc_delta(sum_withdrawals, previous.get("sum_withdrawals"))
        dynamics["sum_transfers"] = calc_delta(sum_transfers, previous.get("sum_transfers"))
        dynamics["total_transactions"] = calc_delta(total_transactions, previous.get("total_transactions"))

    return {
        "week_start": week_start_date.isoformat(),
        "week_end": week_end_date.isoformat(),
        "new_users": new_users,
        "deposit_users": deposit_users,
        "transaction_users": transaction_users,
        "sum_deposits": sum_deposits,
        "sum_withdrawals": sum_withdrawals,
        "sum_transfers": sum_transfers,
        "total_transactions": total_transactions,
        "completed_transactions": completed_transactions,
        "conversions": conversions,
        "avg_deposit": avg_deposit,
        "avg_withdrawal": avg_withdrawal,
        "active_users": active_users,
        "dynamics": dynamics,
    }


async def collect_all_weeks_report() -> Tuple[str, bytes]:
    """
    Collects a report for the last 52 weeks, starting with the current (or previous) week.
    """
    async with async_session_maker() as session:
        report = []
        today = datetime.utcnow().date()
        # last_monday: find the Monday of the current week
        last_monday = today - timedelta(days=today.weekday())
        # start_date: the Monday 52 weeks ago
        start_date = last_monday - timedelta(weeks=52)

        previous_metrics = None
        for i in range(52):
            week_start = datetime.combine(start_date + timedelta(weeks=i), datetime.min.time())
            week_end = week_start + timedelta(days=6)
            metrics = await collect_week_metrics(session, week_start, week_end, previous_metrics)
            report.append(metrics)
            previous_metrics = metrics

    report_json = json.dumps(report, ensure_ascii=False)
    excel_bytes = generate_excel_file(report)

    return report_json, excel_bytes


def generate_excel_file(report_data: List[Dict[str, Any]]) -> bytes:
    """
    Generates an Excel file with three sheets
    """
    wb = Workbook()

    # Sheet 1: Weekly Report
    ws_main = wb.active
    ws_main.title = "Weekly Report"
    main_headers = [
        "week_start",
        "week_end",
        "new_users",
        "deposit_users",
        "transaction_users",
        "sum_deposits",
        "sum_withdrawals",
        "sum_transfers",
        "total_transactions",
        "completed_transactions",
        "avg_deposit",
        "avg_withdrawal",
        "active_users",
    ]
    ws_main.append(main_headers)
    for week in report_data:
        row = [week.get(key) for key in main_headers]
        ws_main.append(row)
    _auto_adjust_column_width(ws_main)

    # Sheet 2: Conversions
    ws_conv = wb.create_sheet(title="Conversions")
    conv_headers = ["week_start", "week_end", "direction", "count", "sum_amount"]
    ws_conv.append(conv_headers)
    for week in report_data:
        week_start = week.get("week_start")
        week_end = week.get("week_end")
        conversions = week.get("conversions") or {}
        if conversions:
            for direction, data in conversions.items():
                direction_str = direction.lower().replace("_", "-")
                row = [
                    week_start,
                    week_end,
                    direction_str,
                    data.get("count", 0),
                    data.get("sum_amount", 0),
                ]
                ws_conv.append(row)
        else:
            ws_conv.append([week_start, week_end, "No conversions", "", ""])
    _auto_adjust_column_width(ws_conv)

    # Sheet 3: Dynamics
    ws_dyn = wb.create_sheet(title="Dynamics")
    dyn_headers = ["week_start", "week_end", "metric", "delta", "pct_change"]
    ws_dyn.append(dyn_headers)
    for week in report_data:
        week_start = week.get("week_start")
        week_end = week.get("week_end")
        dynamics = week.get("dynamics") or {}
        if dynamics:
            for metric, d in dynamics.items():
                row = [
                    week_start,
                    week_end,
                    metric,
                    d.get("delta"),
                    d.get("pct_change"),
                ]
                ws_dyn.append(row)
        else:
            ws_dyn.append([week_start, week_end, "No dynamics", "", ""])
    _auto_adjust_column_width(ws_dyn)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def convert_report_to_json(report_data: List[Dict[str, Any]]) -> str:
    """
    Converts a list of weekly report dictionaries into a JSON string (utf-8).
    """
    return json.dumps(report_data, ensure_ascii=False)


def _auto_adjust_column_width(ws):
    """
    Automatically adjusts the column width to fit the maximum content length.
    """
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

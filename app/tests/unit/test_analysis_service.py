import asyncio
import json
from datetime import date, datetime, time, timedelta
from decimal import Decimal
from io import BytesIO
from typing import Any

import openpyxl
import pytest

from app.models.db_models import Transaction, User
from app.schemas.enums import (CurrencyEnum, TransactionStatusEnum,
                               TransactionTypeEnum)
from app.services import analysis_service
from app.tasks import create_report
from app.tests.utils_test import uniq_email


def _utc(days: int = 0) -> datetime:
    """дней назад от сегодня, время 00:00."""
    d = date.today() - timedelta(days=days)
    return datetime.combine(d, time.min)


async def _seed_data(session):
    """Создаёт 2 пользователей и набор транзакций за последние 7 дней."""
    u1 = User(email=uniq_email("user1"), password="x", created=_utc(5))
    u2 = User(email=uniq_email("user2"), password="x", created=_utc(2))
    session.add_all([u1, u2])
    await session.flush()  # чтобы id появились

    txs: list[Transaction] = [
        # депозиты обоих
        Transaction(
            sender_id=u1.id,
            currency=CurrencyEnum.USD,
            amount=Decimal(100),
            type=TransactionTypeEnum.DEPOSIT,
            status=TransactionStatusEnum.PROCESSED,
            created=_utc(4),
        ),
        Transaction(
            sender_id=u2.id,
            currency=CurrencyEnum.USD,
            amount=Decimal(50),
            type=TransactionTypeEnum.DEPOSIT,
            status=TransactionStatusEnum.PROCESSED,
            created=_utc(1),
        ),
        # вывод u1
        Transaction(
            sender_id=u1.id,
            currency=CurrencyEnum.USD,
            amount=Decimal(40),
            type=TransactionTypeEnum.WITHDRAWAL,
            status=TransactionStatusEnum.PROCESSED,
            created=_utc(3),
        ),
        # обмен u1 USD to EUR
        Transaction(
            sender_id=u1.id,
            recipient_id=u1.id,
            currency=CurrencyEnum.USD,
            amount=Decimal(30),
            type=TransactionTypeEnum.EXCHANGE,
            from_currency=CurrencyEnum.USD,
            to_currency=CurrencyEnum.EUR,
            status=TransactionStatusEnum.PROCESSED,
            created=_utc(2),
        ),
    ]
    session.add_all(txs)
    await session.commit()


@pytest.mark.asyncio
async def test_helper_functions(db_session):
    await _seed_data(db_session)

    start, end = _utc(6).date(), _utc().date()

    assert await analysis_service.get_new_users_count(db_session, start, end) == 2

    assert await analysis_service.get_distinct_senders_count(db_session, start, end) == 2
    assert (
        await analysis_service.get_distinct_senders_count(db_session, start, end, txn_type=TransactionTypeEnum.DEPOSIT)
        == 2
    )

    # суммы
    assert (
        await analysis_service.get_transaction_sum(
            db_session,
            start,
            end,
            txn_type=TransactionTypeEnum.DEPOSIT,
            txn_status=TransactionStatusEnum.PROCESSED,
        )
        == 150.0
    )
    # средняя величина вывода
    assert (
        await analysis_service.get_average_transaction_amount(
            db_session,
            start,
            end,
            txn_type=TransactionTypeEnum.WITHDRAWAL,
            txn_status=TransactionStatusEnum.PROCESSED,
        )
        == 40.0
    )

    # конвертация
    conv = await analysis_service.get_conversions(db_session, start, end)
    assert conv == {"USD_to_EUR": {"count": 1, "sum_amount": 30.0}}

    # подсчёт транзакций
    assert await analysis_service.get_transaction_count(db_session, start, end) == 4
    assert (
        await analysis_service.get_transaction_count(db_session, start, end, txn_status=TransactionStatusEnum.PROCESSED)
        == 4
    )


@pytest.mark.asyncio
async def test_collect_week_metrics(db_session):
    """
    Проверяем, что collect_week_metrics возвращает корректные значения
    и правильно высчитывает динамику по сравнению с «нулевой» предыдущей
    неделей.
    """
    # наполняем БД тестовыми пользователями/транзакциями
    await _seed_data(db_session)

    week_start = _utc(6)  # понедельник «той» недели
    week_end = week_start + timedelta(days=6)

    metrics = await analysis_service.collect_week_metrics(db_session, week_start, week_end, previous=None)

    # минимальный набор ключей
    for key in (
        "new_users",
        "sum_deposits",
        "sum_withdrawals",
        "conversions",
        "dynamics",
    ):
        assert key in metrics

    # создание эталона
    exp_new_users = await analysis_service.get_new_users_count(db_session, week_start.date(), week_end.date())
    exp_sum_deposits = await analysis_service.get_transaction_sum(
        db_session,
        week_start.date(),
        week_end.date(),
        txn_type=TransactionTypeEnum.DEPOSIT,
        txn_status=TransactionStatusEnum.PROCESSED,
    )
    exp_sum_withdrawals = await analysis_service.get_transaction_sum(
        db_session,
        week_start.date(),
        week_end.date(),
        txn_type=TransactionTypeEnum.WITHDRAWAL,
        txn_status=TransactionStatusEnum.PROCESSED,
    )
    exp_conv = await analysis_service.get_conversions(db_session, week_start.date(), week_end.date())

    assert metrics["new_users"] == exp_new_users
    assert metrics["sum_deposits"] == exp_sum_deposits
    assert metrics["sum_withdrawals"] == exp_sum_withdrawals
    assert metrics["conversions"] == exp_conv

    # проверка динамики
    empty_prev = {
        "new_users": 0,
        "sum_deposits": 0,
        "sum_withdrawals": 0,
        "sum_transfers": 0,
        "total_transactions": 0,
    }
    metrics_next = await analysis_service.collect_week_metrics(db_session, week_start, week_end, previous=empty_prev)

    assert metrics_next["dynamics"]["new_users"]["delta"] == exp_new_users
    # деление на ноль pct_change == None
    assert metrics_next["dynamics"]["new_users"]["pct_change"] is None


def _open_wb(raw: bytes):
    return openpyxl.load_workbook(filename=BytesIO(raw))


def test_generate_excel_and_json():
    # упрощённый «отчёт» из двух недель
    fake_report = [
        {
            "week_start": "2025-04-07",
            "week_end": "2025-04-13",
            "new_users": 3,
            "deposit_users": 2,
            "transaction_users": 3,
            "sum_deposits": 120.0,
            "sum_withdrawals": 20.0,
            "sum_transfers": 0.0,
            "total_transactions": 5,
            "completed_transactions": 5,
            "conversions": {"usd_to_eur": {"count": 1, "sum_amount": 30}},
            "avg_deposit": 60.0,
            "avg_withdrawal": 20.0,
            "active_users": 3,
            "dynamics": {},
        },
        {
            "week_start": "2025-03-31",
            "week_end": "2025-04-06",
            "new_users": 1,
            "deposit_users": 1,
            "transaction_users": 1,
            "sum_deposits": 50.0,
            "sum_withdrawals": 0.0,
            "sum_transfers": 0.0,
            "total_transactions": 1,
            "completed_transactions": 1,
            "conversions": {},
            "avg_deposit": 50.0,
            "avg_withdrawal": 0.0,
            "active_users": 1,
            "dynamics": {},
        },
    ]

    # Excel
    raw = analysis_service.generate_excel_file(fake_report)
    wb = _open_wb(raw)
    assert wb.sheetnames == ["Weekly Report", "Conversions", "Dynamics"]
    assert wb["Weekly Report"].max_row == 3  # заголовок + 2 строки

    # JSON
    js = analysis_service.convert_report_to_json(fake_report)
    parsed = json.loads(js)
    assert parsed[0]["new_users"] == 3


def test_generate_weekly_report(monkeypatch):
    """
    Подменяем collect_all_weeks_report и Redis‑клиент, проверяем,
    что таска отдаёт True и кладёт данные в кеш.
    """

    fake_json = "[{}]"
    fake_excel = b"file"

    async def _fake_collect():
        return fake_json, fake_excel

    monkeypatch.setattr(create_report, "collect_all_weeks_report", _fake_collect, raising=True)
    monkeypatch.setattr("asyncio.run", lambda coro: asyncio.get_event_loop().run_until_complete(coro))
    stored: dict[str, Any] = {}

    class _DummyRedis:
        def setex(self, key, ttl, value):
            stored[key] = (ttl, value)

    monkeypatch.setattr(create_report, "redis_cache", _DummyRedis(), raising=True)

    assert create_report.generate_weekly_report() is True
    assert stored["weekly_report_json"][1] == fake_json
    assert stored["weekly_report_excel"][1] == fake_excel
